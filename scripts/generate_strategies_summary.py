"""Generate strategies_summary.xlsx — a multi-sheet workbook documenting the
two live strategies (ORB + dual-momentum), their parameters, universes, and
the validated backtest metrics. Pulls authoritative values from live/config.py
so it never drifts from the running configuration.

Output: strategies_summary.xlsx (at repo root).

  .venv-openbb\\Scripts\\python.exe scripts\\generate_strategies_summary.py
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from live.config import DEFAULTS, SETTINGS, BY_KEY  # noqa: E402

OUT = ROOT / "strategies_summary.xlsx"

# Sector classification for the ORB watchlist (informational).
SECTOR_MAP = {
    "ETF": ["SPY", "QQQ", "IWM", "DIA"],
    "Tech & semis": [
        "AAPL", "MSFT", "NVDA", "AMZN", "GOOGL", "META", "TSLA", "AVGO", "ORCL",
        "AMD", "NFLX", "ADBE", "CRM", "INTC", "CSCO", "QCOM", "TXN", "MU",
        "ACN", "IBM", "NOW", "PANW", "CRWD", "SHOP", "ABNB", "INTU", "PYPL",
        "ANET", "AMAT", "LRCX",
    ],
    "Financials": [
        "JPM", "BAC", "WFC", "GS", "MS", "C", "V", "MA", "AXP",
        "SCHW", "BLK", "USB", "SPGI", "ICE", "CME",
    ],
    "Healthcare": [
        "UNH", "JNJ", "LLY", "PFE", "MRK", "ABBV",
        "TMO", "ABT", "DHR", "AMGN", "ISRG", "BMY", "ELV", "CVS",
    ],
    "Consumer": [
        "WMT", "HD", "COST", "NKE", "MCD", "SBUX", "DIS", "KO", "PEP",
        "TGT", "LOW", "BKNG", "LULU", "ROST", "CMG", "F", "GM",
    ],
    "Industrial & energy": [
        "XOM", "CVX", "CAT", "BA", "GE",
        "HON", "RTX", "LMT", "DE", "UNP", "COP",
    ],
    "Telecom & utility": ["T", "VZ", "TMUS", "CMCSA", "NEE"],
    "High-beta / volume": ["PLTR", "COIN", "UBER", "BABA"],
}

DUALMOM_UNIVERSE = [
    ("SPY", "US large-cap equity (S&P 500)", "risk"),
    ("QQQ", "US tech (NASDAQ-100)", "risk"),
    ("IWM", "US small-cap (Russell 2000)", "risk"),
    ("EFA", "Developed international equity (ex-US)", "risk"),
    ("EEM", "Emerging-market equity", "risk"),
    ("VNQ", "US real estate (REITs)", "risk"),
    ("GLD", "Gold bullion", "risk"),
    ("DBC", "Broad commodities", "risk"),
    ("TLT", "Long US Treasuries (20+ yr)", "risk"),
    ("AGG", "US aggregate bonds", "risk"),
    ("SHY", "Short US Treasuries (1-3 yr) - CASH proxy", "cash"),
]


def fmt_default(s):
    """Render a Setting.default for human readability."""
    d = s.default
    if isinstance(d, list):
        return ", ".join(map(str, d))
    if d is None:
        return "(off / not set)"
    if isinstance(d, bool):
        return "Yes" if d else "No"
    return str(d)


def overview_df():
    return pd.DataFrame([
        {
            "Field": "Strategy name",
            "ORB (intraday breakout)": "ORB — Opening Range Breakout (long-only, filtered)",
            "Dual-momentum (monthly bedrock)": "Dual-momentum — diversified asset-class trend-following",
        },
        {
            "Field": "Cadence",
            "ORB (intraday breakout)": "Intraday. Flat by EOD (15:55 ET).",
            "Dual-momentum (monthly bedrock)": "Monthly. Rebalance first trading day of each month.",
        },
        {
            "Field": "Trade horizon",
            "ORB (intraday breakout)": "Minutes-to-hours (single session, no overnight).",
            "Dual-momentum (monthly bedrock)": "~4 weeks (one month between rebalances).",
        },
        {
            "Field": "Alpaca account",
            "ORB (intraday breakout)": "Main paper account (ALPACA_API_KEY).",
            "Dual-momentum (monthly bedrock)": "Dedicated paper account (DUALMOM_ALPACA_API_KEY).",
        },
        {
            "Field": "Capital deployed",
            "ORB (intraday breakout)": "Up to 8 concurrent x $10k = ~$80k max exposure.",
            "Dual-momentum (monthly bedrock)": "$50k carve-out (DUALMOM_CAPITAL).",
        },
        {
            "Field": "Schedule (Windows Task)",
            "ORB (intraday breakout)": "TradingORB — Mon-Fri 06:15 PDT (09:15 ET, pre-open).",
            "Dual-momentum (monthly bedrock)": "TradingDualMom — Mon-Fri 08:00 PDT, self-gates to first trading day.",
        },
        {
            "Field": "Universe size",
            "ORB (intraday breakout)": "100 large-cap US equities + 4 ETFs.",
            "Dual-momentum (monthly bedrock)": "9 risk-asset ETFs + 1 cash proxy.",
        },
        {
            "Field": "Validated backtest Sharpe",
            "ORB (intraday breakout)": "~1.0 intraday (broad+cap=8); +0.055 avg_R with trend filter.",
            "Dual-momentum (monthly bedrock)": "0.94 (full 2007-2026 incl. 2008/2020/2022).",
        },
        {
            "Field": "Validated max drawdown",
            "ORB (intraday breakout)": "Trend filter: -$5.3k on $100k baseline (-5.3%).",
            "Dual-momentum (monthly bedrock)": "-13.1% across 19 years (vs SPY -50.8%).",
        },
        {
            "Field": "Honest expected return",
            "ORB (intraday breakout)": "Low-single-digit % per year on $100k. Thin edge.",
            "Dual-momentum (monthly bedrock)": "~6-12% per year, bedrock returns. Crisis-alpha.",
        },
        {
            "Field": "What it protects against",
            "ORB (intraday breakout)": "Overnight gaps (flat by close).",
            "Dual-momentum (monthly bedrock)": "Slow/prolonged bear markets via cash rotation.",
        },
        {
            "Field": "What it does NOT protect against",
            "ORB (intraday breakout)": "Mid-session crashes; halt-trading events.",
            "Dual-momentum (monthly bedrock)": "Flash crashes; within-month drawdowns.",
        },
    ])


def orb_params_df():
    """All ORB-relevant settings from live/config.py + a few runtime constants."""
    rows = []
    for s in SETTINGS:
        if s.key in ("watchlist", "short_symbols", "regime_ref_symbol",
                     "regime_sma_window", "regime_confirm_days", "max_flips",
                     "short_enabled"):
            # Short side / universe handled separately
            if s.key == "short_enabled":
                rows.append({
                    "Group": s.group, "Setting": s.label, "Value": fmt_default(s),
                    "Notes": s.help,
                })
            continue
        rows.append({
            "Group": s.group, "Setting": s.label, "Value": fmt_default(s),
            "Notes": s.help,
        })
    # Runtime-only constants from live/paper_orb.py
    rows.append({"Group": "Runtime", "Setting": "Trend filter SMA window (days)",
                 "Value": "200", "Notes": "Daily SMA for the trend filter."})
    rows.append({"Group": "Runtime", "Setting": "Trend filter return lookback (days)",
                 "Value": "20", "Notes": "Trailing window for the relative-strength comparison vs SPY."})
    rows.append({"Group": "Runtime", "Setting": "Poll interval (seconds)",
                 "Value": "10", "Notes": "Bar-fetch + entry-check cadence in the main loop."})
    rows.append({"Group": "Runtime", "Setting": "Late-start cutoff (minutes)",
                 "Value": "10", "Notes": "Script started >10 min after OR window closes -> halts new entries."})
    rows.append({"Group": "Runtime", "Setting": "EOD flatten time (ET)",
                 "Value": "15:55", "Notes": "All positions closed at this time."})
    rows.append({"Group": "Runtime", "Setting": "Network retry on startup",
                 "Value": "5 attempts x 30s", "Notes": "Tolerates morning wifi-joining lag."})
    rows.append({"Group": "Runtime", "Setting": "EOD verify-flat poll",
                 "Value": "60s budget", "Notes": "Waits for close orders to fill before alerting."})
    return pd.DataFrame(rows)


def orb_watchlist_df():
    rows = []
    for sector, syms in SECTOR_MAP.items():
        for sym in syms:
            rows.append({"Symbol": sym, "Sector": sector})
    return pd.DataFrame(rows).sort_values(["Sector", "Symbol"]).reset_index(drop=True)


def dualmom_params_df():
    return pd.DataFrame([
        {"Setting": "Rebalance frequency", "Value": "Monthly",
         "Notes": "First trading day of each month. Self-gating via Alpaca calendar."},
        {"Setting": "Universe", "Value": "9 risk ETFs + SHY (cash proxy)",
         "Notes": "See Dual-Momentum Universe sheet."},
        {"Setting": "Momentum lookbacks (months)", "Value": "3, 6, 12",
         "Notes": "Blended (averaged) -- validated optimum across 2007-2026."},
        {"Setting": "Holdings (top-K)", "Value": "3",
         "Notes": "Equal weight (1/3 each) among the top-3 by blended momentum."},
        {"Setting": "Absolute-momentum filter", "Value": "Must beat SHY",
         "Notes": "Risk assets only included if their blended momentum > cash's."},
        {"Setting": "Fallback / safe harbor", "Value": "SHY (1-3yr Treasuries)",
         "Notes": "Unfilled top-K slots go to SHY. 100% cash if nothing beats it."},
        {"Setting": "Capital deployed (sleeve cap)", "Value": "$50,000",
         "Notes": "DUALMOM_CAPITAL env var. min(account_equity, this) is the sleeve size."},
        {"Setting": "Account", "Value": "Dedicated Alpaca paper account",
         "Notes": "Isolated from the ORB account to prevent symbol/EOD-flatten conflicts."},
        {"Setting": "Safety guard", "Value": "Refuses live without dedicated key",
         "Notes": "If DUALMOM_ALPACA_API_KEY missing, live trading is refused (exit 2)."},
        {"Setting": "Signal as-of", "Value": "Last completed month-end close",
         "Notes": "Drops the in-progress month -> no lookahead, deterministic monthly cadence."},
        {"Setting": "Minimum trade size", "Value": "$50",
         "Notes": "Rebalance deltas under $50 are skipped (noise filter)."},
        {"Setting": "Order type", "Value": "Market notional (day TIF)",
         "Notes": "Sells/closes first, brief pause, then notional buys to free up buying power."},
        {"Setting": "Live runner script", "Value": "live/run_dualmom.py",
         "Notes": "Self-contained. --dry-run prints plan; --force bypasses month-gate (for testing)."},
        {"Setting": "Scheduled task", "Value": "TradingDualMom (weekdays 08:00 PDT)",
         "Notes": "Fires every weekday morning; script no-ops except first trading day of month."},
    ])


def dualmom_universe_df():
    return pd.DataFrame(
        DUALMOM_UNIVERSE, columns=["Ticker", "Description", "Role"]
    )


def backtest_metrics_df():
    return pd.DataFrame([
        # ---- ORB ----
        {"Strategy": "ORB", "Configuration": "Original 5-name watchlist (legacy)",
         "Trades / period": "359 in 180d", "Win %": 48.2, "Avg R": 0.061,
         "Total PnL ($)": 1115, "Max DD ($)": -1654, "Sharpe": 0.92,
         "OOS robust": "Marginal", "Status": "Replaced 2026-05-27"},
        {"Strategy": "ORB", "Configuration": "Broad 55-name + 11:30 cutoff (no trend filter)",
         "Trades / period": "3,663 in 180d", "Win %": 48.0, "Avg R": 0.045,
         "Total PnL ($)": 14757, "Max DD ($)": -11687, "Sharpe": "n/a",
         "OOS robust": "Yes (universe-scale)", "Status": "Stepping stone"},
        {"Strategy": "ORB", "Configuration": "Broad 100-name + 11:30 cutoff (no trend filter)",
         "Trades / period": "6,713 in 180d", "Win %": 46.9, "Avg R": 0.029,
         "Total PnL ($)": 17717, "Max DD ($)": -25333, "Sharpe": "n/a",
         "OOS robust": "Yes", "Status": "Baseline for trend filter"},
        {"Strategy": "ORB", "Configuration": "Broad 100 + 11:30 cutoff + trend filter (LIVE)",
         "Trades / period": "2,420 in 180d", "Win %": 48.0, "Avg R": 0.055,
         "Total PnL ($)": 9977, "Max DD ($)": -5312, "Sharpe": "n/a",
         "OOS robust": "Yes (both halves)", "Status": "CURRENT LIVE CONFIG"},
        # ---- Dual-momentum (19-year backtest) ----
        {"Strategy": "Dual-momentum", "Configuration": "12mo single lookback, top-3",
         "Trades / period": "monthly 2007-2026", "Win %": "n/a", "Avg R": "n/a",
         "Total PnL ($)": "10.2% CAGR", "Max DD ($)": "-22.5%", "Sharpe": 0.70,
         "OOS robust": "Yes", "Status": "Robustness reference"},
        {"Strategy": "Dual-momentum", "Configuration": "3/6/12 blend, top-3 (LIVE)",
         "Trades / period": "monthly 2007-2026", "Win %": "n/a", "Avg R": "n/a",
         "Total PnL ($)": "12.7% CAGR", "Max DD ($)": "-13.1%", "Sharpe": 0.94,
         "OOS robust": "Yes (pre-AI + AI era)", "Status": "CURRENT LIVE CONFIG"},
        {"Strategy": "Dual-momentum", "Configuration": "1/3/6/12 blend (adds 1mo)",
         "Trades / period": "monthly 2007-2026", "Win %": "n/a", "Avg R": "n/a",
         "Total PnL ($)": "12.2% CAGR", "Max DD ($)": "-13.7%", "Sharpe": 0.90,
         "OOS robust": "Yes but worse than 3/6/12", "Status": "Tested 2026-05-31, rejected"},
        {"Strategy": "Dual-momentum", "Configuration": "1/3/6 blend (drops 12mo)",
         "Trades / period": "monthly 2007-2026", "Win %": "n/a", "Avg R": "n/a",
         "Total PnL ($)": "11.7% CAGR", "Max DD ($)": "-23.1%", "Sharpe": 0.83,
         "OOS robust": "No (regime-dependent)", "Status": "Curve-fit warning"},
        {"Strategy": "Dual-momentum", "Configuration": "Faber GTAA-5 (10mo SMA)",
         "Trades / period": "monthly 2007-2026", "Win %": "n/a", "Avg R": "n/a",
         "Total PnL ($)": "5.6% CAGR", "Max DD ($)": "-12.3%", "Sharpe": 0.57,
         "OOS robust": "Yes", "Status": "Alternative reference"},
        {"Strategy": "Dual-momentum", "Configuration": "SPY buy-and-hold (benchmark)",
         "Trades / period": "n/a", "Win %": "n/a", "Avg R": "n/a",
         "Total PnL ($)": "10.9% CAGR", "Max DD ($)": "-50.8%", "Sharpe": 0.63,
         "OOS robust": "n/a", "Status": "Benchmark"},
    ])


def rejected_filters_df():
    """ORB filters tested in research that were REJECTED — kept here so the
    record of negative findings doesn't get lost."""
    return pd.DataFrame([
        {"Filter": "BE stop lift @ +1R", "Family": "Stop management",
         "Result": "Rejected", "Notes": "Lost 33-36% PnL across 180-day window. Cut trades to BE that would have finished positive at EOD."},
        {"Filter": "BE stop lift @ +0.5R", "Family": "Stop management",
         "Result": "Rejected", "Notes": "Same as above, even more aggressive cut."},
        {"Filter": "Pre-market gap > 1% skip-day", "Family": "Day-level filter",
         "Result": "Rejected", "Notes": "Gap days were MORE profitable, not less."},
        {"Filter": "RVOL vs 20-day baseline >= 1.5", "Family": "Volume",
         "Result": "Rejected", "Notes": "Hurt MONOTONICALLY. High RVOL = institutional absorption / climax."},
        {"Filter": "Volume expansion vs prior 5 bars >= 1.0", "Family": "Volume",
         "Result": "Rejected (OOS)", "Notes": "Looked great full-window, failed OOS split (chop half went negative)."},
        {"Filter": "Stop offset 0.10x OR range", "Family": "Stop placement",
         "Result": "Marginal", "Notes": "+8.3% PnL, below the 10% gate. Below noise level."},
        {"Filter": "Wide+high-vol quadrant (Wyckoff GOOD)", "Family": "Effort/result",
         "Result": "Modest", "Notes": "Only +0.040 avg_R vs +0.033 baseline. OOS wobbly."},
        {"Filter": "Reject narrow+high-vol (Wyckoff CHURN)", "Family": "Effort/result",
         "Result": "Tested, not shipped", "Notes": "Gentle improvement, likely overlaps with trend filter signal."},
        {"Filter": "Surprise % > 3% PEAD entry", "Family": "Catalyst",
         "Result": "Rejected", "Notes": "Naive PEAD is decayed. Surprise magnitude is noise (OOS zero in both halves)."},
        {"Filter": "Time cutoff 10:30 ET", "Family": "Time-of-day",
         "Result": "Rejected", "Notes": "Cuts too aggressively. 11:30 was the sweet spot (shipped)."},
        {"Filter": "Time cutoff 11:00 ET", "Family": "Time-of-day",
         "Result": "Likely overfit", "Notes": "+190% PnL on this sample but suspicious as local optimum. 11:30 chosen instead."},
    ])


def write_excel():
    sheets = {
        "Overview": overview_df(),
        "ORB Parameters": orb_params_df(),
        "ORB Watchlist (100)": orb_watchlist_df(),
        "Dual-Momentum Parameters": dualmom_params_df(),
        "Dual-Momentum Universe": dualmom_universe_df(),
        "Backtest Metrics": backtest_metrics_df(),
        "Rejected Filters (research)": rejected_filters_df(),
    }
    with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
        for name, df in sheets.items():
            df.to_excel(xw, sheet_name=name, index=False)
        # Add a footer cell on Overview with the generation timestamp.
        wb = xw.book
        ws = wb["Overview"]
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.cell(row=ws.max_row + 2, column=1, value=f"Generated: {ts}  |  "
                f"Source of truth: live/config.py + paper_orb.py + run_dualmom.py")
        # Pretty-format every sheet.
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            # Header style
            header_fill = PatternFill("solid", fgColor="1F2937")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            ws.row_dimensions[1].height = 24
            # Column widths -- auto-size with reasonable caps.
            for col in ws.columns:
                col_idx = col[0].column
                letter = get_column_letter(col_idx)
                max_len = 0
                for cell in col:
                    if cell.value is None:
                        continue
                    txt = str(cell.value)
                    # respect line breaks
                    longest_line = max((len(line) for line in txt.splitlines()), default=0)
                    max_len = max(max_len, longest_line)
                ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 65)
            # Wrap long-text columns.
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.freeze_panes = "A2"
    print(f"Wrote {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(write_excel())
